import argparse
import os
from datetime import datetime
from openpyxl import load_workbook


def prepare(template_path, output_path, state_fullname, end_year):
    """
    Load an .xlsm template, inject full state name and end-year parameters, and save to output.
    """
    wb = load_workbook(template_path, keep_vba=True)
    ws = wb["Main - Scenario Options"]  # adjust sheet name as needed
    ws["D10"] = state_fullname  # full state name cell
    ws["D12"] = end_year        # end-year cell
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="Generate state-specific scenario .xlsm files from templates"
    )
    parser.add_argument(
        "--templates-dir", default="input_scenarios",
        help="Folder containing baseline.xlsm & policy.xlsm"
    )
    parser.add_argument(
        "--output-dir", default="input_scenarios",
        help="Where to write the generated files"
    )
    parser.add_argument(
        "--states-file", default="states.csv",
        help="Path to CSV of state codes and full names (abbr,fullname per line)"
    )
    parser.add_argument(
        "--end-year", type=int, required=True,
        help="Last year of model run (e.g. 2030 or 2050)"
    )
    args = parser.parse_args()

    # Read list of (abbr, fullname) pairs
    states = []
    with open(args.states_file) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            abbr, fullname = line.split(",", 1)
            states.append((abbr.strip(), fullname.strip()))

    for abbr, fullname in states:
        for scenario in ("baseline", "policy"):
            tpl = os.path.join(args.templates_dir, f"{scenario}.xlsm")
            out = os.path.join(
                args.output_dir,
                f"{scenario}_{abbr}_{args.end_year}.xlsm"
            )
            print(f"[{datetime.now()}] Preparing {out} for {fullname}")
            prepare(
                tpl,
                out,
                fullname,
                args.end_year
            )


if __name__ == "__main__":
    main()
